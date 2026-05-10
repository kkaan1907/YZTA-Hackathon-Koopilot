from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from typing import List
import io
import csv
from database import get_db
import models
import schemas

router = APIRouter(
    prefix="/inventory",
    tags=["Inventory"]
)

@router.get("", response_model=List[schemas.ProductResponse], summary="Mevcut ürünleri ve stok durumlarını listele")
def get_inventory(db: Session = Depends(get_db)):
    products = db.query(models.Product).all()
    return products

@router.get("/alerts", response_model=List[schemas.ProductResponse], summary="Stok seviyesi düşük ( < 5) olan ürünleri listele")
def get_inventory_alerts(db: Session = Depends(get_db)):
    products = db.query(models.Product).filter(models.Product.stock < 5).all()
    return products

@router.post("", response_model=schemas.ProductResponse, summary="Yeni ürün ekle")
def create_product(product: schemas.ProductCreate, db: Session = Depends(get_db)):
    db_product = models.Product(
        name=product.name,
        description=product.description,
        category=product.category,
        unit=product.unit,
        stock=product.stock,
        price=product.price
    )
    db.add(db_product)
    db.commit()
    db.refresh(db_product)
    return db_product

@router.put("/{product_id}", response_model=schemas.ProductResponse, summary="Ürün bilgilerini güncelle")
def update_product(product_id: int, update_data: schemas.ProductCreate, db: Session = Depends(get_db)):
    product = db.query(models.Product).filter(models.Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Ürün bulunamadı")
    
    product.name = update_data.name
    product.description = update_data.description
    product.category = update_data.category
    product.unit = update_data.unit
    product.stock = update_data.stock
    product.price = update_data.price
    
    db.commit()
    db.refresh(product)
    return product

@router.post("/upload", summary="Excel veya CSV dosyasından toplu ürün yükle")
async def upload_inventory(file: UploadFile = File(...), db: Session = Depends(get_db)):
    content = await file.read()
    
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Dosya boyutu çok büyük. Maksimum 5MB yüklenebilir.")
        
    try:
        rows = []
        if file.filename.endswith('.xlsx'):
            try:
                from openpyxl import load_workbook
            except ImportError:
                raise HTTPException(status_code=400, detail="XLSX desteği için openpyxl kurulmalıdır. CSV yüklemeyi deneyebilirsiniz.")

            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            sheet = workbook.active
            headers = [str(cell.value).strip().lower() if cell.value is not None else "" for cell in next(sheet.iter_rows(max_row=1))]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                rows.append({headers[i]: row[i] for i in range(min(len(headers), len(row)))})
        elif file.filename.endswith('.csv'):
            decoded = content.decode("utf-8-sig")
            rows = list(csv.DictReader(io.StringIO(decoded)))
            rows = [{str(k).strip().lower(): v for k, v in row.items()} for row in rows]
        else:
            raise HTTPException(status_code=400, detail="Sadece .xlsx veya .csv dosyaları desteklenmektedir.")
            
        if len(rows) > 5000:
            raise HTTPException(status_code=400, detail="Dosya çok büyük. Maksimum 5000 satır yüklenebilir.")
        
        required_cols = ['name', 'category', 'stock', 'price']
        for col in required_cols:
            if not rows or col not in rows[0]:
                raise HTTPException(status_code=400, detail=f"Eksik sütun: {col}")
        
        updates = 0
        creations = 0
        for index, row in enumerate(rows, start=2):
            try:
                stock_val = float(str(row.get('stock') or 0).replace(",", "."))
                price_val = float(str(row.get('price') or 0).replace(",", "."))
                
                if stock_val < 0 or price_val < 0:
                    continue
                    
                product_name = str(row.get('name') or '').strip()
                if not product_name:
                    continue
                product = db.query(models.Product).filter(models.Product.name == product_name).first()
                
                if product:
                    product.stock = stock_val
                    product.price = price_val
                    product.category = str(row.get('category') or '').strip()
                    if row.get('unit'): 
                        product.unit = str(row.get('unit')).strip()
                    if row.get('description'): 
                        product.description = str(row.get('description')).strip()
                    updates += 1
                else:
                    new_prod = models.Product(
                        name=product_name,
                        category=str(row.get('category') or '').strip(),
                        stock=stock_val,
                        price=price_val,
                        unit=str(row.get('unit') or 'Adet').strip(),
                        description=str(row.get('description') or '').strip()
                    )
                    db.add(new_prod)
                    creations += 1
            except Exception as e:
                print(f"Row {index} processing error: {e}")
                continue
        db.commit()
        return {"status": "success", "updates": updates, "creations": creations}
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Dosya işlenirken hata oluştu: {str(e)}")
